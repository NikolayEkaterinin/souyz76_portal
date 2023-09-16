from django.shortcuts import render, redirect
from .models import Objects
from django.views import View
from openpyxl import load_workbook

from .models import Regions, Objects


class UploadExcelView(View):
    template_name = 'objects_loading_excel.html'

    def post(self, request):
        excel_file = request.FILES.get('excel_file')

        if excel_file:
            if excel_file.name.endswith('.xlsx'):
                try:
                    wb = load_workbook(excel_file)
                    sheet = wb.active

                    for row in sheet.iter_rows(values_only=True):
                        name, sap, address, region_name = row[:4]  # Предполагается, что первые четыре столбца соответствуют полям модели

                        # Поиск или создание региона
                        region, created = Regions.objects.get_or_create(name=region_name)

                        obj = Objects(name=name, sap=sap, address=address, regions=region)
                        obj.save()

                    return redirect('object_list')  # Перенаправление на список объектов
                except Exception as e:
                    return render(request, self.template_name, {'error_message': f'Ошибка при обработке файла: {e}'})
            else:
                return render(request, self.template_name, {'error_message': 'Пожалуйста, загрузите файл Excel (.xlsx).'})
        else:
            return render(request, self.template_name, {'error_message': 'Файл Excel не найден.'})

    def get(self, request):
        return render(request, self.template_name)
